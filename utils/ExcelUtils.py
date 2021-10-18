from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from os import path, makedirs

from unidecode import unidecode


class excel():
    def __init__(self, file_path=".\\database\\",
                 file_name=datetime.now().strftime("%d_%m_%Y") + "_metricas", file_extension=".xlsx"):
        '''When you create Instance of excel, you can change file_path, file_name and file_extension'''
        self._wb = None
        self._ws = None
        self._file_path = file_path
        self._file_name = file_name
        self._file_extension = file_extension
        self._dict_datas = {}

    def _configure_header(self):
        font = Font(color="FFFFFF", bold=True, size=12)
        fill = PatternFill("solid", fgColor="357ae8")
        alignment = Alignment(horizontal="center", vertical="center")
        self._ws.cell(row=1, column=1).font = font
        self._ws.cell(row=1, column=2).font = font
        self._ws.cell(row=1, column=3).font = font
        self._ws.cell(row=1, column=4).font = font
        self._ws.cell(row=1, column=5).font = font
        self._ws.cell(row=1, column=1).fill = fill
        self._ws.cell(row=1, column=2).fill = fill
        self._ws.cell(row=1, column=3).fill = fill
        self._ws.cell(row=1, column=4).fill = fill
        self._ws.cell(row=1, column=5).fill = fill
        self._ws.cell(row=1, column=1).alignment = alignment
        self._ws.cell(row=1, column=2).alignment = alignment
        self._ws.cell(row=1, column=3).alignment = alignment
        self._ws.cell(row=1, column=4).alignment = alignment
        self._ws.cell(row=1, column=5).alignment = alignment

    def read_excel(self, db):
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _rows = self._ws.rows
        _count = 1
        for _row in _rows:
            if (_row[0].value is None or 'titulo' in unidecode(_row[2].value.lower())):
                continue
            if (db == True):
                self._dict_datas[_count] = {'id': _row[0].value, 'id_card': _row[1].value, 'titulo': _row[2].value,
                                            'vaga': _row[3].value, 'nivel': _row[4].value, 'motivo_de_recusa': '',
                                            'como_soube_da_vaga': _row[5].value, 'email_candidato': _row[6].value,
                                            'status_card': _row[7].value, 'email_responsavel': _row[8].value,
                                            'data_entrevista': _row[9].value, 'horario_entrevista': _row[10].value,
                                            'link_entrevista': _row[11].value, 'curriculo_candidato': _row[12].value,
                                            'canal_contato': _row[13].value, 'email_cliente': _row[14].value,
                                            'data_criacao': _row[15].value, 'data_modificacao': _row[16].value,
                                            'linkedin': _row[17].value, 'status': _row[18].value}

                _count += 1
        self._wb.close()
        return self._dict_datas

    def get_last_row(self):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _last_row = self._ws.max_row
        return self._ws[_last_row]

    def write_excel(self):
        '''Write Excel receive a dictionary'''
        self._wb = Workbook()
        self._ws = self._wb.active
        self._ws.cell(row=1, column=1).value = 'Título'
        self._ws.cell(row=1, column=2).value = 'Vaga'
        self._ws.cell(row=1, column=3).value = 'Nível'
        self._ws.cell(row=1, column=4).value = 'Motivo Recusa'
        self._ws.cell(row=1, column=5).value = 'Como soube da vaga?'
        self._configure_header()

        for index, data in self._dict_datas.items():
            self._ws.cell(row=int(index) + 1, column=1).value = data['titulo']
            self._ws.cell(row=int(index) + 1, column=2).value = data['vaga']
            self._ws.cell(row=int(index) + 1, column=3).value = data['nivel']
            self._ws.cell(row=int(index) + 1, column=4).value = data['motivo_recusa']
            self._ws.cell(row=int(index) + 1, column=5).value = data['soube_vaga']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._ws.title = 'Dashboard'
        self._wb.save(self._file_path + self._file_name + "_2" + self._file_extension)
        return self._file_path + self._file_name + self._file_extension

    def update_excel(self, dict_datas):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _max_row = self._ws.max_row

        if (len(dict_datas.keys()) > 5):
            self._ws.title = 'Report'
            self._ws.cell(row=int(_max_row), column=1).value = _max_row
            self._ws.cell(row=int(_max_row), column=2).value = self._dict_datas['id_card']
            self._ws.cell(row=int(_max_row), column=3).value = self._dict_datas['titulo']
            self._ws.cell(row=int(_max_row), column=4).value = self._dict_datas['vaga']
            self._ws.cell(row=int(_max_row), column=5).value = self._dict_datas['nivel']
            self._ws.cell(row=int(_max_row), column=6).value = self._dict_datas['motivo_recusa']
            self._ws.cell(row=int(_max_row), column=7).value = self._dict_datas['soube_vaga']
        else:
            self._ws.title = 'Dashboard'
            for index, data in self._dict_datas.items():
                self._ws.cell(row=int(index) + _max_row, column=1).value = data['titulo']
                self._ws.cell(row=int(index) + _max_row, column=2).value = data['vaga']
                self._ws.cell(row=int(index) + _max_row, column=3).value = data['nivel']
                self._ws.cell(row=int(index) + _max_row, column=4).value = data['motivo_recusa']
                self._ws.cell(row=int(index) + _max_row, column=5).value = data['soube_vaga']

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._wb.save(self._file_path + self._file_name + self._file_extension)
        self._wb.close()
        return self._file_path + self._file_name + self._file_extension

    def update_id(self, dict_datas):
        self._dict_datas = dict_datas
        self._wb = load_workbook(self._file_path + self._file_name + self._file_extension)
        self._ws = self._wb.active
        _rows = self._ws.rows
        _count = 1
        for _row in _rows:
            if (_row[1].value == dict_datas['id_card']):
                if (self._dict_datas['titulo']):
                    self._ws.cell(row=_count, column=1).value = self._dict_datas['titulo']
                if (self._dict_datas['vaga']):
                    self._ws.cell(row=_count, column=2).value = self._dict_datas['vaga']
                if (self._dict_datas['nivel']):
                    self._ws.cell(row=_count, column=3).value = self._dict_datas['nivel']
                if (self._dict_datas['motivo_recusa']):
                    self._ws.cell(row=_count, column=4).value = self._dict_datas['motivo_recusa']
                if (self._dict_datas['como_soube_da_vaga']):
                    self._ws.cell(row=_count, column=5).value = self._dict_datas['como_soube_da_vaga']
                if (self._dict_datas['email_candidato']):
                    self._ws.cell(row=_count, column=6).value = self._dict_datas['email_candidato']
                if (self._dict_datas['status_card']):
                    self._ws.cell(row=_count, column=7).value = self._dict_datas['status_card']
                if (self._dict_datas['email_responsavel']):
                    self._ws.cell(row=_count, column=8).value = self._dict_datas['email_responsavel']
                if (self._dict_datas['data_entrevista']):
                    self._ws.cell(row=_count, column=9).value = self._dict_datas['data_entrevista']
                if (self._dict_datas['horario_entrevista']):
                    self._ws.cell(row=_count, column=10).value = self._dict_datas['horario_entrevista']
                if (self._dict_datas['link_entrevista']):
                    self._ws.cell(row=_count, column=11).value = self._dict_datas['link_entrevista']
                if (self._dict_datas['curriculo_candidato']):
                    self._ws.cell(row=_count, column=12).value = self._dict_datas['curriculo_candidato']
                if (self._dict_datas['canal_contato']):
                    self._ws.cell(row=_count, column=13).value = self._dict_datas['canal_contato']
                if (self._dict_datas['email_cliente']):
                    self._ws.cell(row=_count, column=14).value = self._dict_datas['email_cliente']
                if (self._dict_datas['data_criacao']):
                    self._ws.cell(row=_count, column=15).value = self._dict_datas['data_criacao']
                if (self._dict_datas['data_modificacao']):
                    self._ws.cell(row=_count, column=16).value = self._dict_datas['data_modificacao']
                if (self._dict_datas['linkedin']):
                    self._ws.cell(row=_count, column=17).value = self._dict_datas['linkedin']
                if (self._dict_datas['status']):
                    self._ws.cell(row=_count, column=18).value = self._dict_datas['status']
            else:
                _count += 1

        # Save the file
        if (not path.exists(self._file_path)):
            makedirs(self._file_path)

        self._wb.save(self._file_path + self._file_name + self._file_extension)
        self._wb.close()
        return self._file_path + self._file_name + self._file_extension
